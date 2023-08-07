from bitcoin import compress, privkey_to_pubkey, pubkey_to_address, der_encode_sig, ecdsa_raw_verify, der_decode_sig, N, \
    decode_privkey, fast_multiply, hash_to_int, inv, G, hashlib, safe_from_hex, deterministic_generate_k
from blockcypher import create_unsigned_tx, verify_unsigned_tx, broadcast_signed_transaction, get_transaction_details
from blockcypher.constants import COIN_SYMBOL_MAPPINGS
from blockcypher.utils import is_valid_coin_symbol

from loguru import logger
from schemes.cht04.btc.settings import SK, MONITORING_ADDRESS, BTC_SYMBOL, PREFERENCE, BTC_API_TOKEN, RECEIVE_ADDRESS
from utils.btc.tools import dec_to_hex
from utils.crypto.crypto import big_small_end_convert
from utils.crypto.crypto import to_even
from utils.crypto.crypto import add_to_len
from utils.crypto.crypto import base58decode_to_P2PKH


def ecdsa_raw_sign(msghash, priv, k):
    logger.info("msghash:" + msghash)
    z = hash_to_int(msghash)
    logger.info(z)
    # k = deterministic_generate_k(msghash, priv)

    k = 2 * k + 1  # 这里等式右边的k应该要限制不超过254比特，所以左边的k不超过255比特，肯定小于N，不需要对N求余
    # k = 10000000000
    r, y = fast_multiply(G, k)
    s = inv(k, N) * (z + r * decode_privkey(priv)) % N

    return 27 + ((y % 2) ^ (0 if s * 2 < N else 1)), r, s if s * 2 < N else N - s


def make_tx_signatures(txs_to_sign, privkey_list, pubkey_list, k=10000000000):
    """
    Loops through txs_to_sign and makes signatures using privkey_list and pubkey_list

    Not sure what privkeys and pubkeys to supply?
    Use get_input_addresses() to return a list of addresses.
    Matching those addresses to keys is up to you and how you store your private keys.
    A future version of this library may handle this for you, but it is not trivial.

    Note that if spending multisig funds the process is significantly more complicated.
    Each tx_to_sign must be signed by *each* private key.
    In a 2-of-3 transaction, two of [privkey1, privkey2, privkey3] must sign each tx_to_sign

    http://dev.blockcypher.com/#multisig-transactions
    """
    assert len(privkey_list) == len(pubkey_list) == len(txs_to_sign)
    # in the event of multiple inputs using the same pub/privkey,
    # that privkey should be included multiple times

    signatures = []
    for cnt, tx_to_sign in enumerate(txs_to_sign):
        sig = der_encode_sig(*ecdsa_raw_sign(tx_to_sign.rstrip(' \t\r\n\0'), privkey_list[cnt], k))
        err_msg = 'Bad Signature: sig %s for tx %s with pubkey %s' % (
            sig,
            tx_to_sign,
            pubkey_list[cnt],
        )
        assert ecdsa_raw_verify(tx_to_sign, der_decode_sig(sig), pubkey_list[cnt]), err_msg
        signatures.append(sig)
    return signatures


def simple_spend(from_privkey=SK, k=10000000000, to_address=RECEIVE_ADDRESS, to_satoshis=10, change_address=None,
                 privkey_is_compressed=False, min_confirmations=0, api_key=BTC_API_TOKEN, coin_symbol=BTC_SYMBOL,
                 preference=PREFERENCE):
    '''
    Simple method to spend from one single-key address to another.

    Signature takes place locally (client-side) after unsigned transaction is verified.

    Returns the tx_hash of the newly broadcast tx.

    If no change_address specified, change will be sent back to sender address.
    Note that this violates the best practice.

    To sweep, set to_satoshis=-1

    Compressed public keys (and their corresponding addresses) have been the standard since v0.6,
    set privkey_is_compressed=False if using uncompressed addresses.

    Note that this currently only supports spending from single key addresses.
    '''
    assert is_valid_coin_symbol(coin_symbol), coin_symbol
    assert isinstance(to_satoshis, int), to_satoshis
    assert api_key, 'api_key required'

    if privkey_is_compressed:
        from_pubkey = compress(privkey_to_pubkey(from_privkey))
    else:
        from_privkey = dec_to_hex(from_privkey)
        from_pubkey = privkey_to_pubkey(from_privkey)
    from_address = pubkey_to_address(
        pubkey=from_pubkey,
        # this method only supports paying from pubkey anyway
        magicbyte=COIN_SYMBOL_MAPPINGS[coin_symbol]['vbyte_pubkey'],
    )

    inputs = [{'address': from_address}, ]
    logger.info('inputs: %s' % inputs)
    outputs = [{'address': to_address, 'value': to_satoshis}, ]
    logger.info('outputs: %s' % outputs)

    # will fail loudly if tx doesn't verify client-side
    unsigned_tx = create_unsigned_tx(
        inputs=inputs,
        outputs=outputs,
        # may build with no change address, but if so will verify change in next step
        # done for extra security in case of client-side bug in change address generation
        change_address=change_address,
        coin_symbol=coin_symbol,
        min_confirmations=min_confirmations,
        verify_tosigntx=False,  # will verify in next step
        include_tosigntx=True,
        api_key=api_key,
        preference=preference,
    )
    logger.info('unsigned_tx: %s' % unsigned_tx)

    if 'errors' in unsigned_tx:
        print('TX Error(s): Tx NOT Signed or Broadcast')
        for error in unsigned_tx['errors']:
            print(error['error'])
        # Abandon
        raise Exception('Build Unsigned TX Error')

    if change_address:
        change_address_to_use = change_address
    else:
        change_address_to_use = from_address

    tx_is_correct, err_msg = verify_unsigned_tx(
        unsigned_tx=unsigned_tx,
        inputs=inputs,
        outputs=outputs,
        sweep_funds=bool(to_satoshis == -1),
        change_address=change_address_to_use,
        coin_symbol=coin_symbol,
    )
    if not tx_is_correct:
        print(unsigned_tx)  # for debug
        raise Exception('TX Verification Error: %s' % err_msg)

    privkey_list, pubkey_list = [], []
    for proposed_input in unsigned_tx['tx']['inputs']:
        privkey_list.append(from_privkey)
        pubkey_list.append(from_pubkey)
        # paying from a single key should only mean one address per input:
        assert len(proposed_input['addresses']) == 1, proposed_input['addresses']
    # logger.info('privkey_list: %s' % privkey_list)
    logger.info('pubkey_list: %s' % pubkey_list)

    # sign locally
    tx_signatures = make_tx_signatures(
        txs_to_sign=unsigned_tx['tosign'],
        privkey_list=privkey_list,
        pubkey_list=pubkey_list,
        k=k
    )
    logger.info('tx_signatures: %s' % tx_signatures)

    # broadcast TX
    # broadcasted_tx='errors'
    broadcasted_tx = broadcast_signed_transaction(
        unsigned_tx=unsigned_tx,
        signatures=tx_signatures,
        pubkeys=pubkey_list,
        coin_symbol=coin_symbol,
        api_key=api_key,
    )
    logger.info('broadcasted_tx: %s' % broadcasted_tx)

    if 'errors' in broadcasted_tx:
        print('TX Error(s): Tx May NOT Have Been Broadcast')
        for error in broadcasted_tx['errors']:
            print(error['error'])
        print(broadcasted_tx)
        return

    return broadcasted_tx['tx']['hash']


def simulate_spend(from_privkey=SK, k=10000000000, to_address=RECEIVE_ADDRESS, to_satoshis=10, change_address=None,
                   privkey_is_compressed=False, min_confirmations=0, api_key=BTC_API_TOKEN, coin_symbol=BTC_SYMBOL,
                   preference=PREFERENCE):
    """
    模拟发送交易，取消最后的广播过程
    Simple method to spend from one single-key address to another.

    Signature takes place locally (client-side) after unsigned transaction is verified.

    Returns the tx_hash of the newly broadcast tx.

    If no change_address specified, change will be sent back to sender address.
    Note that this violates the best practice.

    To sweep, set to_satoshis=-1

    Compressed public keys (and their corresponding addresses) have been the standard since v0.6,
    set privkey_is_compressed=False if using uncompressed addresses.

    Note that this currently only supports spending from single key addresses.
    """
    assert is_valid_coin_symbol(coin_symbol), coin_symbol
    assert isinstance(to_satoshis, int), to_satoshis
    assert api_key, 'api_key required'

    if privkey_is_compressed:
        from_pubkey = compress(privkey_to_pubkey(from_privkey))
    else:
        from_privkey = dec_to_hex(from_privkey)
        from_pubkey = privkey_to_pubkey(from_privkey)
    from_address = pubkey_to_address(
        pubkey=from_pubkey,
        # this method only supports paying from pubkey anyway
        magicbyte=COIN_SYMBOL_MAPPINGS[coin_symbol]['vbyte_pubkey'],
    )

    inputs = [{'address': from_address}, ]
    logger.info('inputs: %s' % inputs)
    outputs = [{'address': to_address, 'value': to_satoshis}, ]
    logger.info('outputs: %s' % outputs)

    # will fail loudly if tx doesn't verify client-side
    unsigned_tx = create_unsigned_tx(
        inputs=inputs,
        outputs=outputs,
        # may build with no change address, but if so will verify change in next step
        # done for extra security in case of client-side bug in change address generation
        change_address=change_address,
        coin_symbol=coin_symbol,
        min_confirmations=min_confirmations,
        verify_tosigntx=False,  # will verify in next step
        include_tosigntx=True,
        api_key=api_key,
        preference=preference,
    )
    logger.info('unsigned_tx: %s' % unsigned_tx)

    if 'errors' in unsigned_tx:
        print('TX Error(s): Tx NOT Signed or Broadcast')
        for error in unsigned_tx['errors']:
            print(error['error'])
        # Abandon
        raise Exception('Build Unsigned TX Error')

    if change_address:
        change_address_to_use = change_address
    else:
        change_address_to_use = from_address

    tx_is_correct, err_msg = verify_unsigned_tx(
        unsigned_tx=unsigned_tx,
        inputs=inputs,
        outputs=outputs,
        sweep_funds=bool(to_satoshis == -1),
        change_address=change_address_to_use,
        coin_symbol=coin_symbol,
    )
    if not tx_is_correct:
        print(unsigned_tx)  # for debug
        raise Exception('TX Verification Error: %s' % err_msg)

    privkey_list, pubkey_list = [], []
    for proposed_input in unsigned_tx['tx']['inputs']:
        privkey_list.append(from_privkey)
        pubkey_list.append(from_pubkey)
        # paying from a single key should only mean one address per input:
        assert len(proposed_input['addresses']) == 1, proposed_input['addresses']
    # logger.info('privkey_list: %s' % privkey_list)
    logger.info('pubkey_list: %s' % pubkey_list)

    # sign locally
    tx_signatures = make_tx_signatures(
        txs_to_sign=unsigned_tx['tosign'],
        privkey_list=privkey_list,
        pubkey_list=pubkey_list,
        k=k
    )
    logger.info('tx_signatures: %s' % tx_signatures)

    # broadcast TX
    # broadcasted_tx='errors'
    # broadcasted_tx = broadcast_signed_transaction(
    #     unsigned_tx=unsigned_tx,
    #     signatures=tx_signatures,
    #     pubkeys=pubkey_list,
    #     coin_symbol=coin_symbol,
    #     api_key=api_key,
    # )
    # logger.info('broadcasted_tx: %s' % broadcasted_tx)
    #
    # if 'errors' in broadcasted_tx:
    #     print('TX Error(s): Tx May NOT Have Been Broadcast')
    #     for error in broadcasted_tx['errors']:
    #         print(error['error'])
    #     print(broadcasted_tx)
    #     return
    #
    # return broadcasted_tx['tx']['hash']
    return 1


def get_k_by_tx_hash(tx_hash, sk):
    tx_info = get_transaction_details(tx_hash=tx_hash, coin_symbol=BTC_SYMBOL, api_key=BTC_API_TOKEN, include_hex=True)
    # print(tx_info)
    # 用索引为0的输入计算k
    msghash = get_msghash_from_txinfo(tx_info)[0]
    logger.info("msghash:" + msghash)
    sig_length_hex_string = tx_info['inputs'][0]['script'][:2]
    sig_length_dec = int(sig_length_hex_string, 16) * 2
    sig = tx_info['inputs'][0]['script'][2:sig_length_dec]
    logger.info("sig:" + sig)
    z = hash_to_int(msghash)
    logger.info("z:" + str(z))
    v, r, s = der_decode_sig(sig)
    logger.info("r and s: " + str(r) + ' ' + str(s))

    k = inv(s, N) * (z + r * decode_privkey(sk)) % N
    logger.info("k:" + str(k))

    if k % 2 == 0:
        k = N - k

    k = (k - 1) * inv(2, N) % N

    return k


# 暂时只支持输入的锁定脚本为P2PKH类型的交易,且交易输入和输出数量不超过ff(隐蔽传输的角度足够了)
def get_msghash_from_txinfo(tx_info):
    input_num = int(tx_info['vin_sz'])
    logger.info("input_num:" + str(input_num))
    output_num = int(tx_info['vout_sz'])
    logger.info("output_num:" + str(output_num))

    hex_string = tx_info['hex']
    logger.info("hex_string:" + str(hex_string))
    begin = 8 + 2
    input_struct_begin_flag_list = list()
    input_struct_begin_flag_list.append(begin)
    # 计算各签名起始长度位置
    for i in range(input_num):
        if i == 0:
            continue
        else:
            last_input_sig_length_hex_string = hex_string[input_struct_begin_flag_list[i - 1] + 64 + 8:
                                                          input_struct_begin_flag_list[i - 1] + 64 + 8 + 2]
            last_input_sig_length_dec = int(last_input_sig_length_hex_string, 16) * 2
            input_struct_begin_flag_list.append(
                input_struct_begin_flag_list[i - 1] + 64 + 8 + 2 + last_input_sig_length_dec + 8)

    logger.info("input_struct_begin_flag_list:" + str(input_struct_begin_flag_list))

    final_input_struct_begin_flag = input_struct_begin_flag_list[input_num - 1]
    logger.info("final_input_struct_begin_flag:" + str(final_input_struct_begin_flag))
    final_input_sig_length_hex_string = hex_string[
                                        final_input_struct_begin_flag + 64 + 8: final_input_struct_begin_flag + 64 + 8 + 2]
    logger.info("final_input_sig_length_hex_string:" + str(final_input_sig_length_hex_string))
    final_input_sig_length_dec = int(final_input_sig_length_hex_string, 16) * 2
    logger.info("final_input_sig_length_dec:" + str(final_input_sig_length_dec))
    input_struct_end_flag = final_input_struct_begin_flag + 64 + 8 + 2 + final_input_sig_length_dec + 8
    logger.info("input_struct_end_flag:" + str(input_struct_end_flag))
    input_struct_begin_flag_list.append(input_struct_end_flag)
    logger.info("input_struct_begin_flag_list:" + str(input_struct_begin_flag_list))
    # output_struct_begin_falg = input_struct_end_flag + 2
    # output_value_begin_flag_list = list()
    # output_script_or_opreturn_begin_flag_list = list()
    # # 计算各输出金额起始位置 和 script或者opreturn起始位置
    # for i in range(output_num):
    #     if i == 0:
    #         output_value_begin_flag_list.append(output_struct_begin_falg)
    #     else:
    #         last_script_or_opreturn_length_hex_string = hex_string[output_script_or_opreturn_begin_flag_list[i - 1]: output_script_or_opreturn_begin_flag_list[i - 1] + 2]
    #         last_value_and_script_or_opreturn_length_dec = int(last_script_or_opreturn_length_hex_string,
    #                                                            16) * 2
    #         output_value_begin_flag_list.append(
    #             output_value_begin_flag_list[i - 1] + 16 + 2 + last_value_and_script_or_opreturn_length_dec)
    #     output_script_or_opreturn_begin_flag_list.append(output_value_begin_flag_list[i] + 16)

    # 开始构建tosign_tx
    tosign_tx_list = list()
    for i in range(input_num):
        version = hex_string[:8]
        logger.info("version:" + str(version))
        input_number = add_to_len(big_small_end_convert(to_even(hex(input_num)[2:]).encode()), 2).decode()
        logger.info("input_number:" + str(input_number))
        input_struct = ''
        for j in range(input_num):
            if i == j:
                hash_and_index = hex_string[input_struct_begin_flag_list[j]: input_struct_begin_flag_list[j] + 72]
                logger.info("hash_and_index:" + str(hash_and_index))
                script_type = tx_info['inputs'][i]['script_type']
                try:
                    if script_type != 'pay-to-pubkey-hash':
                        raise Exception(f'脚本类型（{script_type}）不是pay-to-pubkey-hash！暂时无法解析！')
                except Exception:
                    logger.exception(f'脚本类型（{script_type}）不是pay-to-pubkey-hash！暂时无法解析！')
                    raise
                pkScript = base58decode_to_P2PKH(tx_info['inputs'][i]['addresses'][0])
                logger.info("pkScript:" + str(pkScript))
                sequence = hex_string[input_struct_begin_flag_list[j + 1] - 8: input_struct_begin_flag_list[j + 1]]
                logger.info("sequence:" + str(sequence))
                input_struct += hash_and_index + pkScript + sequence
            else:
                hash_and_index = hex_string[input_struct_begin_flag_list[j]: input_struct_begin_flag_list[j] + 72]
                logger.info("hash_and_index:" + str(hash_and_index))
                sequence = hex_string[input_struct_begin_flag_list[j + 1] - 8: input_struct_begin_flag_list[j + 1]]
                logger.info("sequence:" + str(sequence))
                input_struct += hash_and_index + '00' + sequence

        # output_number = add_to_len(big_small_end_convert(to_even(hex(output_num)[2:]).encode()), 2).decode()
        # logger.info("output_number:" + str(output_number))
        output_struct = hex_string[input_struct_begin_flag_list[input_num]:]
        logger.info("output_struct:" + str(output_struct))
        hash_type = '01000000'  # 这个东西说是hash类型，不知道具体指什么的哈希,可能是输入的脚本签名中的hashtype，不一定
        tosign_tx_list.append(version + input_number + input_struct + output_struct + hash_type)

    logger.info("tosign_tx_list:" + str(tosign_tx_list))

    msghash_list = list()
    for i in range(input_num):
        msghash = hashlib.sha256(hashlib.sha256(safe_from_hex(tosign_tx_list[i])).digest()).hexdigest()
        msghash_list.append(msghash)
    logger.info("msghash_list:" + str(msghash_list))

    return msghash_list


def get_msghash_from_hex(tx_hex):
    head = tx_hex[:82]
    padding = tx_hex[-60:-8]
    middle = tx_hex[-154:-60]
    tail = '0000000001000000'
    tosign = head + padding + middle + padding + tail
    logger.info("tosign:" + tosign)
    # print(tosign)
    msghash = hashlib.sha256(hashlib.sha256(safe_from_hex(tosign)).digest()).hexdigest()
    logger.info("msghash:" + msghash)
    return msghash


if __name__ == '__main__':
    # # 昊天你可以在这里测试 k不超过254比特就行

    # # 这个sk因为我要用我自己的地址测试所以把它提出来了
    # sk = 17010883408893843693166803031195714567951202172038459263433240037419908661000

    # # 发交易的时候change_address也可以选择,可以试试，msghash也不会错
    # print(simple_spend(k=19996554910799302231190538777410434851253670, from_privkey=sk, to_satoshis=22222, to_address='muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz')) # 嵌入k

    """
    2020-07-16 01:02:56.602 | INFO     | __main__:simple_spend:97 - inputs: [{'address': 'mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'}]
    2020-07-16 01:02:56.602 | INFO     | __main__:simple_spend:99 - outputs: [{'address': 'muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz', 'value': 22222}]
    2020-07-16 01:02:57.910 | INFO     | __main__:simple_spend:115 - unsigned_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': '5ef1f0e02885cc3da92333a8d2882ea7c22b40fccc4082b33374ebc0e4d9943c', 'addresses': ['mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9', 'muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz'], 'total': 951910, 'fees': 3500, 'size': 119, 'preference': 'medium', 'relayed_by': '103.116.47.196', 'received': '2020-07-15T17:02:59.462882977Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '6ff2ba43f8bd8a8b7dd3d9218a4edf4548a321107a3c97742146564829835775', 'output_index': 2, 'output_value': 955410, 'sequence': 4294967295, 'addresses': ['mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780610}], 'outputs': [{'value': 22222, 'script': '76a914978c4313f44902c08ce344844a726b61c04655b688ac', 'addresses': ['muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 929688, 'script': '76a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac', 'addresses': ['mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign_tx': ['0100000001755783294856462174973c7a1021a34845df4e8a21d9d37d8b8abdf843baf26f020000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488acffffffff02ce560000000000001976a914978c4313f44902c08ce344844a726b61c04655b688ac982f0e00000000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac0000000001000000'], 'tosign': ['249e6d8041fdc7d3f926b208609d8115d26ed42386391eefea9199948a409056']}
    2020-07-16 01:02:57.911 | INFO     | __main__:simple_spend:148 - pubkey_list: ['045826f20f985779e1a20a151f9be4045831718ad6cf8db7c459a3e0aa23430d263279baca1855aa69d87be60e99a68d64e2ba862fd061649f2c854f45704c911b']
    2020-07-16 01:02:57.911 | INFO     | __main__:ecdsa_raw_sign:17 - msghash:249e6d8041fdc7d3f926b208609d8115d26ed42386391eefea9199948a409056
    2020-07-16 01:02:57.911 | INFO     | __main__:ecdsa_raw_sign:19 - 16563180133407246505203780522585941388572593974610370320390640640140408557654
    2020-07-16 01:02:57.922 | INFO     | __main__:simple_spend:157 - tx_signatures: ['3045022100dfe391303f88230ff159575a6b37ccb8fa3864b86abb8afb7875ecc371839b1f022070debb0a2395362d56aea8de2b60a1054ef535e62b468179d5ddf07582cc7a62']
    e32cba41d36bb7ea3efed9e84492148244610233b1449615242c57d4a9b0d0d4
    2020-07-16 01:02:59.419 | INFO     | __main__:simple_spend:168 - broadcasted_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': 'e32cba41d36bb7ea3efed9e84492148244610233b1449615242c57d4a9b0d0d4', 'addresses': ['muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz', 'mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'], 'total': 951910, 'fees': 3500, 'size': 258, 'preference': 'low', 'relayed_by': '103.116.47.196', 'received': '2020-07-15T17:03:00.693390122Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '6ff2ba43f8bd8a8b7dd3d9218a4edf4548a321107a3c97742146564829835775', 'output_index': 2, 'script': '483045022100dfe391303f88230ff159575a6b37ccb8fa3864b86abb8afb7875ecc371839b1f022070debb0a2395362d56aea8de2b60a1054ef535e62b468179d5ddf07582cc7a620141045826f20f985779e1a20a151f9be4045831718ad6cf8db7c459a3e0aa23430d263279baca1855aa69d87be60e99a68d64e2ba862fd061649f2c854f45704c911b', 'output_value': 955410, 'sequence': 4294967295, 'addresses': ['mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780610}], 'outputs': [{'value': 22222, 'script': '76a914978c4313f44902c08ce344844a726b61c04655b688ac', 'addresses': ['muLGMbZMDi9iQ48BLBqeNQ1qk7xBf1E7bz'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 929688, 'script': '76a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac', 'addresses': ['mpuecvTPQPTreyWDwaZ84b7SBQMkv3y3Z9'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign': ['']}
    
    
    """
    # 恢复的时候需要用到私钥，我也提出来了
    # print(get_k_by_tx_hash('e32cba41d36bb7ea3efed9e84492148244610233b1449615242c57d4a9b0d0d4', sk=sk)) # 提取k
    """
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:209 - input_num:1
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:211 - output_num:2
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:214 - hex_string:0100000001755783294856462174973c7a1021a34845df4e8a21d9d37d8b8abdf843baf26f020000008b483045022100dfe391303f88230ff159575a6b37ccb8fa3864b86abb8afb7875ecc371839b1f022070debb0a2395362d56aea8de2b60a1054ef535e62b468179d5ddf07582cc7a620141045826f20f985779e1a20a151f9be4045831718ad6cf8db7c459a3e0aa23430d263279baca1855aa69d87be60e99a68d64e2ba862fd061649f2c854f45704c911bffffffff02ce560000000000001976a914978c4313f44902c08ce344844a726b61c04655b688ac982f0e00000000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac00000000
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:227 - input_struct_begin_flag_list:[10]
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:230 - final_input_struct_begin_flag:10
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:232 - final_input_sig_length_hex_string:8b
    2020-07-16 01:03:44.514 | INFO     | __main__:get_msghash_from_txinfo:234 - final_input_sig_length_dec:278
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:236 - input_struct_end_flag:370
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:238 - input_struct_begin_flag_list:[10, 370]
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:258 - version:01000000
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:260 - input_number:01
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:265 - hash_and_index:755783294856462174973c7a1021a34845df4e8a21d9d37d8b8abdf843baf26f02000000
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:274 - pkScript:1976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:276 - sequence:ffffffff
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:288 - output_struct:02ce560000000000001976a914978c4313f44902c08ce344844a726b61c04655b688ac982f0e00000000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac00000000
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:292 - tosign_tx_list:['0100000001755783294856462174973c7a1021a34845df4e8a21d9d37d8b8abdf843baf26f020000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488acffffffff02ce560000000000001976a914978c4313f44902c08ce344844a726b61c04655b688ac982f0e00000000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac0000000001000000']
    2020-07-16 01:03:44.515 | INFO     | __main__:get_msghash_from_txinfo:298 - msghash_list:['249e6d8041fdc7d3f926b208609d8115d26ed42386391eefea9199948a409056']
    2020-07-16 01:03:44.515 | INFO     | __main__:get_k_by_tx_hash:185 - msghash:249e6d8041fdc7d3f926b208609d8115d26ed42386391eefea9199948a409056
    2020-07-16 01:03:44.515 | INFO     | __main__:get_k_by_tx_hash:189 - sig:3045022100dfe391303f88230ff159575a6b37ccb8fa3864b86abb8afb7875ecc371839b1f022070debb0a2395362d56aea8de2b60a1054ef535e62b468179d5ddf07582cc7a62
    2020-07-16 01:03:44.517 | INFO     | __main__:get_k_by_tx_hash:191 - z:16563180133407246505203780522585941388572593974610370320390640640140408557654
    2020-07-16 01:03:44.517 | INFO     | __main__:get_k_by_tx_hash:193 - r and s: 101267841571761500378020543296295743528343794466037236500258063575426521078559 51052569989620265638323573704057004406755242070218396784986568656852118960738
    2020-07-16 01:03:44.517 | INFO     | __main__:get_k_by_tx_hash:196 - k:115792089237316195423570985008687867859727742680470442001527608320648458986996
    19996554910799302365953195927915915020075008
    """

    """
    2020-07-14 12:24:28.819 | INFO     | __main__:get_k_by_tx_hash:191 - k:115792089237316195423570985008687907852837564279074904382605163141508161494337
    2020-07-14 12:24:28.826 | INFO     | __main__:get_k_by_tx_hash:194 - r:51134462373623003832568098972913963716165477285712924082047273136071059725873
    """

    """
    01000000012337c92b78a9f41ae0f68763138cb3713f140e0a4688a8c785ed19f062acb246010000008a4730440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e3102204a0fb26978b71f5f6dc040088dc999d4f485ce00323ae530dd1c7c13c9b05050014104ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412ffffffff0240e20100000000001976a914eee1596774020e24a1434d9c83e472667d9e17f788ace8360d00000000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac00000000
    2020-07-12 17:15:21.728 | INFO     | __main__:simple_spend:92 - inputs: [{'address': 'mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'}]
    2020-07-12 17:15:21.728 | INFO     | __main__:simple_spend:94 - outputs: [{'address': 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo', 'value': 10}]
    2020-07-12 17:15:24.061 | INFO     | __main__:simple_spend:110 - unsigned_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': 'f5ac96d80cdff1f7ea0cf8acac0f0072c7ea5875e63780b05f2d976a3d5241bd', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 996500, 'fees': 3500, 'size': 119, 'preference': 'medium', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T09:15:24.052561061Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': 'af5424b431e1ce1ca6d889491624237efe3d9805d9917818f9bcafce20f7c790', 'output_index': 1, 'output_value': 1000000, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780501}], 'outputs': [{'value': 10, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 996490, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign_tx': ['010000000190c7f720ceafbcf9187891d905983dfe7e2324164989d8a61ccee131b42454af010000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288acffffffff020a000000000000001976a914eee1596774020e24a1434d9c83e472667d9e17f788ac8a340f00000000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac0000000001000000'], 'tosign': ['cf3dd0b947631757edb627675a66a3a3e462228b9ffc30c61240499cb1b1ae71']}
    2020-07-12 17:15:24.061 | INFO     | __main__:simple_spend:143 - pubkey_list: ['04ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412']
    2020-07-12 17:15:24.071 | INFO     | __main__:simple_spend:152 - tx_signatures: ['30440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e31022027ae011e6e4189f950282e9f10e479022ed6209d8258210fe296f56369cdcdcc']
    72426e4998bad15d39c0171f42885089bf6aad97e90a517f65dce5a37187a34c
    2020-07-12 17:15:25.188 | INFO     | __main__:simple_spend:162 - broadcasted_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': '72426e4998bad15d39c0171f42885089bf6aad97e90a517f65dce5a37187a34c', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 996500, 'fees': 3500, 'size': 257, 'preference': 'low', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T09:15:25.178208186Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': 'af5424b431e1ce1ca6d889491624237efe3d9805d9917818f9bcafce20f7c790', 'output_index': 1, 'script': '4730440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e31022027ae011e6e4189f950282e9f10e479022ed6209d8258210fe296f56369cdcdcc014104ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412', 'output_value': 1000000, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780501}], 'outputs': [{'value': 10, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 996490, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign': ['']}
    
    2020-07-12 17:38:41.524 | INFO     | __main__:simple_spend:94 - inputs: [{'address': 'mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'}]
    2020-07-12 17:38:41.524 | INFO     | __main__:simple_spend:96 - outputs: [{'address': 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo', 'value': 10}]
    2020-07-12 17:38:42.583 | INFO     | __main__:simple_spend:112 - unsigned_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': 'b81d81757af494baa27c1a2222e03ac0e8a03819aff3db532bb7fa8cac448632', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 992990, 'fees': 3500, 'size': 119, 'preference': 'medium', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T09:38:42.606402119Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '72426e4998bad15d39c0171f42885089bf6aad97e90a517f65dce5a37187a34c', 'output_index': 1, 'output_value': 996490, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780502}], 'outputs': [{'value': 10, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 992980, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 
    'tosign_tx': ['01000000014ca38771a3e5dc657f510ae997ad6abf895088421f17c0395dd1ba98496e4272010000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288acffffffff020a000000000000001976a914eee1596774020e24a1434d9c83e472667d9e17f788acd4260f00000000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac0000000001000000'], 
    'tosign': ['29c7c44f5caf4cc940fb44d100644ddd0b63df3f33fc539600add921e55ebf5b']}
    2020-07-12 17:38:42.583 | INFO     | __main__:simple_spend:145 - pubkey_list: ['04ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412']
    2020-07-12 17:38:42.583 | INFO     | __main__:ecdsa_raw_sign:14 - msghash:29c7c44f5caf4cc940fb44d100644ddd0b63df3f33fc539600add921e55ebf5b
    2020-07-12 17:38:42.584 | INFO     | __main__:ecdsa_raw_sign:16 - 18897784239685423108768466402745810405386266652621798812717689565581938704219
    2020-07-12 17:38:42.593 | INFO     | __main__:simple_spend:154 - tx_signatures: ['30440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e3102200fa354369bff95dd19249e1fccc84a4ff124cb89291869743851d7b8996487b9']
    46b2ac62f019ed85c7a888460a0e143f71b38c136387f6e01af4a9782bc93723
    2020-07-12 17:38:43.655 | INFO     | __main__:simple_spend:165 - broadcasted_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': '46b2ac62f019ed85c7a888460a0e143f71b38c136387f6e01af4a9782bc93723', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 992990, 'fees': 3500, 'size': 257, 'preference': 'low', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T09:38:43.667243716Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '72426e4998bad15d39c0171f42885089bf6aad97e90a517f65dce5a37187a34c', 'output_index': 1, 'script': '4730440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e3102200fa354369bff95dd19249e1fccc84a4ff124cb89291869743851d7b8996487b9014104ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412', 'output_value': 996490, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780502}], 'outputs': [{'value': 10, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 992980, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign': ['']}
    
    2020-07-12 19:06:32.390 | INFO     | __main__:simple_spend:94 - inputs: [{'address': 'mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'}]
    2020-07-12 19:06:32.390 | INFO     | __main__:simple_spend:96 - outputs: [{'address': 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo', 'value': 123456}]
    2020-07-12 19:06:33.433 | INFO     | __main__:simple_spend:112 - unsigned_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': '599a6ccf0a0e92989f2db364c05e8789de7476c30acb724c9590b1f82dbc665f', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 989480, 'fees': 3500, 'size': 119, 'preference': 'medium', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T11:06:33.491654716Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '46b2ac62f019ed85c7a888460a0e143f71b38c136387f6e01af4a9782bc93723', 'output_index': 1, 'output_value': 992980, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780504}], 'outputs': [{'value': 123456, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 866024, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 
    'tosign_tx': ['01000000012337c92b78a9f41ae0f68763138cb3713f140e0a4688a8c785ed19f062acb246010000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288acffffffff0240e20100000000001976a914eee1596774020e24a1434d9c83e472667d9e17f788ace8360d00000000001976a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac0000000001000000'], 
    'tosign': ['23b7d8e928bb57c5b385e8e6eea7c62e737f3f5c697159748e09532ca16efa9c']}
    2020-07-12 19:06:33.433 | INFO     | __main__:simple_spend:145 - pubkey_list: ['04ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412']
    2020-07-12 19:06:33.433 | INFO     | __main__:ecdsa_raw_sign:14 - msghash:23b7d8e928bb57c5b385e8e6eea7c62e737f3f5c697159748e09532ca16efa9c
    2020-07-12 19:06:33.433 | INFO     | __main__:ecdsa_raw_sign:16 - 16155779776436808182655086604927275272776837647619500413983696030990615902876
    2020-07-12 19:06:33.444 | INFO     | __main__:simple_spend:154 - tx_signatures: ['30440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e3102204a0fb26978b71f5f6dc040088dc999d4f485ce00323ae530dd1c7c13c9b05050']
    2270738e5e5a0813634cfb174528b9fb9ffaa76472d7fec09a16d5385bf9230e
    2020-07-12 19:06:34.671 | INFO     | __main__:simple_spend:165 - broadcasted_tx: {'tx': {'block_height': -1, 'block_index': -1, 'hash': '2270738e5e5a0813634cfb174528b9fb9ffaa76472d7fec09a16d5385bf9230e', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX', 'n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'total': 989480, 'fees': 3500, 'size': 257, 'preference': 'low', 'relayed_by': '103.116.47.196', 'received': '2020-07-12T11:06:34.723364174Z', 'ver': 1, 'double_spend': False, 'vin_sz': 1, 'vout_sz': 2, 'confirmations': 0, 'inputs': [{'prev_hash': '46b2ac62f019ed85c7a888460a0e143f71b38c136387f6e01af4a9782bc93723', 'output_index': 1, 'script': '4730440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e3102204a0fb26978b71f5f6dc040088dc999d4f485ce00323ae530dd1c7c13c9b05050014104ea8f5787d3b472b18d0152c5ffeb67225883defe7121af149aa01ec0203e7ad6febd5839cde4771ba37473a6f41f287051522131b18fb39480a048361c7cf412', 'output_value': 992980, 'sequence': 4294967295, 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash', 'age': 1780504}], 'outputs': [{'value': 123456, 'script': '76a914eee1596774020e24a1434d9c83e472667d9e17f788ac', 'addresses': ['n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo'], 'script_type': 'pay-to-pubkey-hash'}, {'value': 866024, 'script': '76a914be2664b82a436f50a92bab39d63db11d0a4eca6288ac', 'addresses': ['mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX'], 'script_type': 'pay-to-pubkey-hash'}]}, 'tosign': ['']}
    """

    # from blockcypher import decodetx
    #
    # result = decodetx('ffffffff0240e20100000000001976a914eee1596774020e24a1434d9c83e472667d9e17f788ac2467070000000000', coin_symbol=BTC_SYMBOL, api_key=BTC_API_TOKEN, )
    # print(result)

    # # 验证s变为相反数就正确了，并查看中间变量
    # msghash = '6638a5314e83099819455b1660fa874dc278beb8aa32f047c38dfe69d0b2e8b9'
    # logger.info("msghash:" + msghash)
    # z = hash_to_int(msghash)
    # logger.info('z:' + str(z))
    # # k = deterministic_generate_k(msghash, priv)
    # k = int(10000000000)
    # r, y = fast_multiply(G, k)
    # logger.info('r:' + str(r))
    # priv = dec_to_hex(SK)
    # logger.info('priv:' + str(priv))
    # s = inv(k, N) * (z + r * decode_privkey(priv)) % N
    # logger.info('s:' + str(s))

    """
    2020-07-13 01:11:36.136 | INFO     | __main__:<module>:256 - msghash:6638a5314e83099819455b1660fa874dc278beb8aa32f047c38dfe69d0b2e8b9
    2020-07-13 01:11:36.136 | INFO     | __main__:<module>:258 - z:46235994108573632711523239026063013134045843724271143573208652329378135599289
    2020-07-13 01:11:36.137 | INFO     | __main__:<module>:262 - r:51134462373623003832568098972913963716165477285712924082047273136071059725873
    2020-07-13 01:11:36.137 | INFO     | __main__:<module>:264 - priv:ece77a1b5e35c7f48a59804c7e997eadea18c81f3594f0aeaaf3ef6bdfb0e841
    2020-07-13 01:11:36.137 | INFO     | __main__:<module>:266 - s:63563435294679656083675226800918621703946365184329262101534750017398787582666
    """

    # # 验证s变为相反数就正确了
    # sig = '30440220710d147f7bbe1d4a85a32e7f4050e94d9b6c4d21a2ab400a835225fde1785e31022073785edc76265dd1a0afdad03081b6875281cf5053a9a77818a49e2a2c659677'
    # logger.info("sig:" + sig)
    # z = hash_to_int(msghash)
    # logger.info("z:" + str(z))
    # v, r, s = der_decode_sig(sig)
    #
    # s = N - s
    # logger.info("r and s: " + str(r) + ' ' + str(s))
    #
    # k = inv(s, N) * (z + r * decode_privkey(SK)) % N
    # logger.info("k:" + str(k))

    # # 验证修改v是否对签名和验证有影响
    # sig = der_encode_sig(*ecdsa_raw_sign(msghash, SK, 12345678))
    # print(sig)
    # # vrs = der_decode_sig(sig)
    # # print(vrs)
    # from_privkey = dec_to_hex(SK)
    # from_pubkey = privkey_to_pubkey(from_privkey)
    # res =  ecdsa_raw_verify(msghash, der_decode_sig(sig), from_pubkey)
    # print(res)

    # print(hashlib.sha256(hashlib.sha256(safe_from_hex('0100000001c43fd2d4274987f4b6b9f8a7c4732d13ba6f616ba433d980062a955e653f687d020000001976a9148e2bd174262d4a1eb67598d3de4aac4f4f148e6188acffffffff030000000000000000136a4c105445535432303230303731333233343639300000000000001976a914978c4313f44902c08ce344844a726b61c04655b688ac12940e00000000001976a9146703f1f1a5282dd1ed1d0fb8fa7f0ce1a9e9cef488ac0000000001000000')).digest()).hexdigest())

    # msghash测试
    # tx_info = get_transaction_details(tx_hash='2b9df8a3d283882c49086c2f3991741f43f140bc8e2fe55c81ac5f1e80f7d56c', coin_symbol=BTC_SYMBOL, api_key=BTC_API_TOKEN, include_hex=True)
    # get_msghash_from_txinfo(tx_info)
    #
    # 9e146e1c2134c15307e07accbb8b96869af045865c096bf1cacd19cc4dd56f9e
    # f227916685026a58da9a5f24254a6afcfe43cc05edf96bbaab131b367d558f3b

    # print(simple_spend())

    # txid 测试
    # raw_tx = '01000000026f07409a400ad9a4897a481e6358e702774854e638e8fabe3e25139b9f2f2d04010000008b483045022048293f0152fc9a7e2b3cd9105592d1a8d34fe6b68366430d6f2f5947f78f0a8b022100b74e57953d9e9f2fbc43cedd382c2f2fd24cd070bcca14c716dbc0defa6b75100141045a20dc9145ca40a0bb81bce2486f384ca9085094389be5046fca5b933202fabae15b7c3f1552ac3b901bd460a01669d4facbf30c94c7b13dae930058db1802a3ffffffff58ba38058809083d70a056f952442a357d7586e802cb82ae64fe613aed2a1711010000008a473044022075efcf6df17971da38539cf737899d9d79060d0bcabc971cae531b423e3576ba022016b9dcb92f83551ea9a0de3f276e4e13a872aff2945005180518e06c1f7efde50141049fc22099a0c8fe2830275532f7b73e3040e5e7207c5eb1f09d33a261a6e4b8573e9fc31d9bd3090536b2bfdde39315cbcad3f29be88450116fbe5896059f880cffffffff0218c16100000000001976a9144cef4512d960ef2dcd6b756153c39ee43e96844088ac95fa4800000000001976a914c57eceb2572c19afeddcfe1b191efbc112daa2a688ac00000000'
    # txid = hashlib.sha256(hashlib.sha256(safe_from_hex(raw_tx)).digest()).hexdigest()
    # print(big_small_end_convert(txid))

    # decode_raw_tx 测试
    # raw_tx = '01000000026f07409a400ad9a4897a481e6358e702774854e638e8fabe3e25139b9f2f2d04010000008b483045022048293f0152fc9a7e2b3cd9105592d1a8d34fe6b68366430d6f2f5947f78f0a8b022100b74e57953d9e9f2fbc43cedd382c2f2fd24cd070bcca14c716dbc0defa6b75100141045a20dc9145ca40a0bb81bce2486f384ca9085094389be5046fca5b933202fabae15b7c3f1552ac3b901bd460a01669d4facbf30c94c7b13dae930058db1802a3ffffffff58ba38058809083d70a056f952442a357d7586e802cb82ae64fe613aed2a1711010000008a473044022075efcf6df17971da38539cf737899d9d79060d0bcabc971cae531b423e3576ba022016b9dcb92f83551ea9a0de3f276e4e13a872aff2945005180518e06c1f7efde50141049fc22099a0c8fe2830275532f7b73e3040e5e7207c5eb1f09d33a261a6e4b8573e9fc31d9bd3090536b2bfdde39315cbcad3f29be88450116fbe5896059f880cffffffff0218c16100000000001976a9144cef4512d960ef2dcd6b756153c39ee43e96844088ac95fa4800000000001976a914c57eceb2572c19afeddcfe1b191efbc112daa2a688ac00000000'
    # from blockcypher.api import decodetx
    # print(decodetx(raw_tx, api_key=BTC_API_TOKEN, coin_symbol='btc'))

    # 模拟发送
    # print(simulate_spend())

    from typing import List, Tuple, Dict
    from bitcoin import privkey_to_address, random_key
    from utils.crypto.crypto import big_small_end_convert
    from utils.crypto.crypto import to_even
    from utils.crypto.crypto import add_to_len
    # from blockcypher import make_tx_signatures as mts
    from os.path import join as pjoin
    from openpyxl import load_workbook
    from utils.crypto.crypto import encrypt, decrypt

    import random
    import os
    import json
    import uuid

    def getSingleTx(in_sk: int, prev_hash: str, output_index: int, out_address_or_opreturn: List[Tuple[int, str, int]],
                    total: int, fees: int, k: int):
        """
        :param k:
        :param in_sk:                   输入地址对应的私钥，默认输入地址只有一个，不用List表示
        :param prev_hash:               前序哈希
        :param output_index:            前序哈希索引
        :param out_address_or_opreturn: 输出地址或者opreturn，一个元组组成的列表，元组有三个元素：
                                        第一位是标识位，0代表是opreturn，1代表是地址；
                                        第二位数输出的地址或者op_return；
                                        第三位是输出金额，op_return时金额必须是0；
                                        @example：[(0, "test", 0), (1, "mkQ7ekkE252jT5xHEmioBhSkeMpTL6U9vb", 123772)]
        :param total:                   总金额，即输入金额，或者输出总金额加上手续费
        :param fees:                    手续费
        :return:
        """
        # 地址、脚本和大小端转换等准备工作
        addresses = list()

        # 主链
        in_address = privkey_to_address(in_sk)
        # 测试链
        # in_address = privkey_to_address(in_sk, "111")

        addresses.append(in_address)

        for output in out_address_or_opreturn:
            if output[0] == 1:
                flag = True
                for address in addresses:
                    if output[1] == address:
                        flag = False
                if flag:
                    addresses.append(output[1])

        # 计算tosignTx和msghash
        tosign_tx, tx_dic = getTosignTx(in_sk, prev_hash, output_index, out_address_or_opreturn)
        tosign = hashlib.sha256(hashlib.sha256(safe_from_hex(tosign_tx)).digest()).hexdigest()
        # print("tosign_tx:" + str(tosign_tx))
        # print("tosign:" + str(tosign))

        # 计算签名
        from_privkey = dec_to_hex(in_sk)
        from_pubkey = privkey_to_pubkey(from_privkey)
        # print("from_pubkey:" + str(from_pubkey))

        privkey_list, pubkey_list, txs_to_sign = [], [], []
        privkey_list.append(from_privkey)
        pubkey_list.append(from_pubkey)
        txs_to_sign.append(tosign)

        # tx_signatures = make_tx_signatures(
        #     txs_to_sign=txs_to_sign,
        #     privkey_list=privkey_list,
        #     pubkey_list=pubkey_list,
        #     k=k
        # )

        tx_signatures = make_tx_signatures(
            txs_to_sign=txs_to_sign,
            privkey_list=privkey_list,
            pubkey_list=pubkey_list,
            k=k,
        )

        tx_signature = tx_signatures[0]
        # print("tx_signature:" + str(tx_signature))

        # 计算raw_tx
        tx_signature_length = hex(int(len(tx_signature) / 2) + 1)[2:]
        from_pubkey_length = hex(int(len(from_pubkey) / 2))[2:]
        total_len = hex(int(tx_signature_length, 16) + int(from_pubkey_length, 16) + 2)[2:]
        input_signature = total_len + tx_signature_length + tx_signature + "01" + from_pubkey_length + str(from_pubkey)
        # print("input_signature:" + str(input_signature))

        raw_tx = tx_dic["version"] + tx_dic["input_num"] + tx_dic["prev_hash"] + tx_dic["prev_hash_index"] + \
                 input_signature + tx_dic["sequence"] + tx_dic["output_num"] + tx_dic["output_script"] + '00000000'
        size = int(len(raw_tx) / 2)
        txid = big_small_end_convert(
            hashlib.sha256(hashlib.sha256(safe_from_hex(raw_tx)).digest()).hexdigest()).decode()
        # print("txid:" + str(txid))

        # 获取字典数据
        inputs = list()
        in_addresses = list()
        in_addresses.append(in_address)
        inputs.append({"prev_hash": prev_hash,
                       "output_index": output_index,
                       "script": tx_signature_length + tx_signature + "01" + from_pubkey_length + str(from_pubkey),
                       "sign": tx_signature + "01",
                       "output_value": total,
                       "sequence": 4294967295,
                       "addressed": in_addresses,
                       "script_type": "pay-to-pubkey-hash",
                       # "age": 1,  # TODO
                       })

        return {"hash": str(txid),
                "addresses": addresses,
                "total": total - fees,
                "fees": fees,
                "size": size,
                "vsize": size,
                "preference": "low",
                "vin_sz": 1,
                "vout_sz": len(out_address_or_opreturn),
                "data_protocol": "unknown",
                "inputs": inputs,
                "outputs": tx_dic["outputs"], }

    def getSingleTxWithoutK258(in_sk: int, prev_hash: str, output_index: int, out_address_or_opreturn: List[Tuple[int, str, int]],
                    total: int, fees: int):
        """
        :param k:
        :param in_sk:                   输入地址对应的私钥，默认输入地址只有一个，不用List表示
        :param prev_hash:               前序哈希
        :param output_index:            前序哈希索引
        :param out_address_or_opreturn: 输出地址或者opreturn，一个元组组成的列表，元组有三个元素：
                                        第一位是标识位，0代表是opreturn，1代表是地址；
                                        第二位数输出的地址或者op_return；
                                        第三位是输出金额，op_return时金额必须是0；
                                        @example：[(0, "test", 0), (1, "mkQ7ekkE252jT5xHEmioBhSkeMpTL6U9vb", 123772)]
        :param total:                   总金额，即输入金额，或者输出总金额加上手续费
        :param fees:                    手续费
        :return:
        """
        # 地址、脚本和大小端转换等准备工作
        addresses = list()

        # 主链
        in_address = privkey_to_address(in_sk)
        # 测试链
        # in_address = privkey_to_address(in_sk, "111")

        addresses.append(in_address)

        for output in out_address_or_opreturn:
            if output[0] == 1:
                flag = True
                for address in addresses:
                    if output[1] == address:
                        flag = False
                if flag:
                    addresses.append(output[1])

        # 计算tosignTx和msghash
        tosign_tx, tx_dic = getTosignTx(in_sk, prev_hash, output_index, out_address_or_opreturn)
        tosign = hashlib.sha256(hashlib.sha256(safe_from_hex(tosign_tx)).digest()).hexdigest()
        # print("tosign_tx:" + str(tosign_tx))
        # print("tosign:" + str(tosign))

        # 计算签名
        from_privkey = dec_to_hex(in_sk)
        from_pubkey = privkey_to_pubkey(from_privkey)
        # print("from_pubkey:" + str(from_pubkey))

        privkey_list, pubkey_list, txs_to_sign = [], [], []
        privkey_list.append(from_privkey)
        pubkey_list.append(from_pubkey)
        txs_to_sign.append(tosign)

        # tx_signatures = make_tx_signatures(
        #     txs_to_sign=txs_to_sign,
        #     privkey_list=privkey_list,
        #     pubkey_list=pubkey_list,
        #     k=k
        # )

        tx_signatures = make_tx_signatures_without_k(
            txs_to_sign=txs_to_sign,
            privkey_list=privkey_list,
            pubkey_list=pubkey_list,
        )

        tx_signature = tx_signatures[0]
        # print("tx_signature:" + str(tx_signature))

        # 计算raw_tx
        tx_signature_length = hex(int(len(tx_signature) / 2) + 1)[2:]
        from_pubkey_length = hex(int(len(from_pubkey) / 2))[2:]
        total_len = hex(int(tx_signature_length, 16) + int(from_pubkey_length, 16) + 2)[2:]
        input_signature = total_len + tx_signature_length + tx_signature + "01" + from_pubkey_length + str(from_pubkey)
        # print("input_signature:" + str(input_signature))

        raw_tx = tx_dic["version"] + tx_dic["input_num"] + tx_dic["prev_hash"] + tx_dic["prev_hash_index"] + \
                 input_signature + tx_dic["sequence"] + tx_dic["output_num"] + tx_dic["output_script"] + '00000000'
        size = int(len(raw_tx) / 2)
        txid = big_small_end_convert(
            hashlib.sha256(hashlib.sha256(safe_from_hex(raw_tx)).digest()).hexdigest()).decode()
        # print("txid:" + str(txid))

        # 获取字典数据
        inputs = list()
        in_addresses = list()
        in_addresses.append(in_address)
        inputs.append({"prev_hash": prev_hash,
                       "output_index": output_index,
                       "script": tx_signature_length + tx_signature + "01" + from_pubkey_length + str(from_pubkey),
                       "sign": tx_signature + "01",
                       "output_value": total,
                       "sequence": 4294967295,
                       "addressed": in_addresses,
                       "script_type": "pay-to-pubkey-hash",
                       # "age": 1,  # TODO
                       })

        return {"hash": str(txid),
                "addresses": addresses,
                "total": total - fees,
                "fees": fees,
                "size": size,
                "vsize": size,
                "preference": "low",
                "vin_sz": 1,
                "vout_sz": len(out_address_or_opreturn),
                "data_protocol": "unknown",
                "inputs": inputs,
                "outputs": tx_dic["outputs"], }

    def getSingleTxWithoutK226(in_sk: int, prev_hash: str, output_index: int, out_address_or_opreturn: List[Tuple[int, str, int]],
                    total: int, fees: int):
        """
        :param k:
        :param in_sk:                   输入地址对应的私钥，默认输入地址只有一个，不用List表示
        :param prev_hash:               前序哈希
        :param output_index:            前序哈希索引
        :param out_address_or_opreturn: 输出地址或者opreturn，一个元组组成的列表，元组有三个元素：
                                        第一位是标识位，0代表是opreturn，1代表是地址；
                                        第二位数输出的地址或者op_return；
                                        第三位是输出金额，op_return时金额必须是0；
                                        @example：[(0, "test", 0), (1, "mkQ7ekkE252jT5xHEmioBhSkeMpTL6U9vb", 123772)]
        :param total:                   总金额，即输入金额，或者输出总金额加上手续费
        :param fees:                    手续费
        :return:
        """
        # 地址、脚本和大小端转换等准备工作
        addresses = list()

        # 主链
        in_address = privkey_to_address(in_sk)
        # 测试链
        # in_address = privkey_to_address(in_sk, "111")

        addresses.append(in_address)

        for output in out_address_or_opreturn:
            if output[0] == 1:
                flag = True
                for address in addresses:
                    if output[1] == address:
                        flag = False
                if flag:
                    addresses.append(output[1])

        # 计算tosignTx和msghash
        tosign_tx, tx_dic = getTosignTx(in_sk, prev_hash, output_index, out_address_or_opreturn)
        tosign = hashlib.sha256(hashlib.sha256(safe_from_hex(tosign_tx)).digest()).hexdigest()
        # print("tosign_tx:" + str(tosign_tx))
        # print("tosign:" + str(tosign))

        # 计算签名
        from_privkey = dec_to_hex(in_sk)
        from_pubkey = privkey_to_pubkey(from_privkey)
        # print("from_pubkey:" + str(from_pubkey))

        privkey_list, pubkey_list, txs_to_sign = [], [], []
        privkey_list.append(from_privkey)
        pubkey_list.append(from_pubkey)
        txs_to_sign.append(tosign)

        # tx_signatures = make_tx_signatures(
        #     txs_to_sign=txs_to_sign,
        #     privkey_list=privkey_list,
        #     pubkey_list=pubkey_list,
        #     k=k
        # )

        tx_signatures = make_tx_signatures_without_k(
            txs_to_sign=txs_to_sign,
            privkey_list=privkey_list,
            pubkey_list=pubkey_list,
        )

        tx_signature = tx_signatures[0]
        # print("tx_signature:" + str(tx_signature))

        # 计算raw_tx
        tx_signature_length = hex(int(len(tx_signature) / 2) + 1)[2:]
        # from_pubkey_length = hex(int(len(from_pubkey) / 2))[2:]
        # total_len = hex(int(tx_signature_length, 16) + int(from_pubkey_length, 16) + 2)[2:]

        flag = (int(from_pubkey[-1:], 16) % 2) + 2
        compress_from_pubkey = "0" + str(flag) + from_pubkey[2:66]
        print(compress_from_pubkey)

        from_pubkey_length = hex(int(len(compress_from_pubkey) / 2))[2:]
        print("from_pubkey_length:", from_pubkey_length)
        total_len = hex(int(tx_signature_length, 16) + int(from_pubkey_length, 16) + 2)[2:]
        print("total_len:", total_len)

        input_signature = total_len + tx_signature_length + tx_signature + "01" + from_pubkey_length + str(compress_from_pubkey)
        print("input_signature:" + str(input_signature))

        raw_tx = tx_dic["version"] + tx_dic["input_num"] + tx_dic["prev_hash"] + tx_dic["prev_hash_index"] + \
                 input_signature + tx_dic["sequence"] + tx_dic["output_num"] + tx_dic["output_script"] + '00000000'
        print("raw_tx:", raw_tx)
        size = int(len(raw_tx) / 2)
        print("size:", size)
        txid = big_small_end_convert(
            hashlib.sha256(hashlib.sha256(safe_from_hex(raw_tx)).digest()).hexdigest()).decode()
        # print("txid:" + str(txid))

        # 获取字典数据
        inputs = list()
        in_addresses = list()
        in_addresses.append(in_address)
        inputs.append({"prev_hash": prev_hash,
                       "output_index": output_index,
                       "script": tx_signature_length + tx_signature + "01" + from_pubkey_length + str(compress_from_pubkey),
                       "sign": tx_signature + "01",
                       "output_value": total,
                       "sequence": 4294967295,
                       "addressed": in_addresses,
                       "script_type": "pay-to-pubkey-hash",
                       # "age": 1,  # TODO
                       })

        return {"hash": str(txid),
                "addresses": addresses,
                "total": total - fees,
                "fees": fees,
                "size": size,
                "vsize": size,
                "preference": "low",
                "vin_sz": 1,
                "vout_sz": len(out_address_or_opreturn),
                "data_protocol": "unknown",
                "inputs": inputs,
                "outputs": tx_dic["outputs"], }

    def getSingleMultiInputTx(in_sk_num: int, out_num: int, total: int, fees: int):
        """
        :param k:
        :param in_sk:                   输入地址对应的私钥，默认输入地址只有一个，不用List表示
        :param prev_hash:               前序哈希
        :param output_index:            前序哈希索引
        :param out_address_or_opreturn: 输出地址或者opreturn，一个元组组成的列表，元组有三个元素：
                                        第一位是标识位，0代表是opreturn，1代表是地址；
                                        第二位数输出的地址或者op_return；
                                        第三位是输出金额，op_return时金额必须是0；
                                        @example：[(0, "test", 0), (1, "mkQ7ekkE252jT5xHEmioBhSkeMpTL6U9vb", 123772)]
        :param total:                   总金额，即输入金额，或者输出总金额加上手续费
        :param fees:                    手续费
        :return:
        """
        # 地址、脚本和大小端转换等准备工作
        addresses = list()

        # 随机生成输入的私钥、对应地址、对应金额和tosign
        in_sk_list = list()
        in_address_list = list()
        txs_to_sign = list()
        in_value_list, prev_hash_list, output_index_list, privkey_list, pubkey_list = [], [], [], [], []
        sum_value = total
        for index in range(in_sk_num):
            prev_hash_list.append(str(big_small_end_convert(geneStrRandomHex64()).decode()))
            output_index_list.append(str(add_to_len(big_small_end_convert(to_even(hex(geneIndex())[2:]).encode()), 8).decode()))
            sk_temp = geneRandomKey()
            from_privkey_temp = dec_to_hex(sk_temp)
            from_pubkey_temp = privkey_to_pubkey(from_privkey_temp)
            privkey_list.append(from_privkey_temp)
            pubkey_list.append(from_pubkey_temp)
            in_sk_list.append(sk_temp)
            # 主链
            in_address_temp = privkey_to_address(sk_temp)
            # 测试链
            # in_address_temp = privkey_to_address(sk_temp, "111")
            in_address_list.append(in_address_temp)
            addresses.append(in_address_temp)
            txs_to_sign.append(geneStrRandomHex64())
            if index < in_sk_num - 1:
                value = random.randint(1, sum_value - in_sk_num + index + 1)
            else:
                value = sum_value
            in_value_list.append(value)
            sum_value = sum_value - value
        print('sum_value:', sum_value)

        # 计算签名
        tx_signatures = make_tx_signatures_without_k(
            txs_to_sign=txs_to_sign,
            privkey_list=privkey_list,
            pubkey_list=pubkey_list,
        )

        # 获取raw_tx的input及其字典数据
        inputs = list()
        raw_inputs = str(add_to_len(big_small_end_convert(to_even(hex(in_sk_num)[2:]).encode()), 2).decode())
        for i in range(in_sk_num):
            tx_signature_length = hex(int(len(tx_signatures[i]) / 2) + 1)[2:]
            from_pubkey_length = hex(int(len(pubkey_list[i]) / 2))[2:]
            total_len = hex(int(tx_signature_length, 16) + int(from_pubkey_length, 16) + 2)[2:]
            input_signature = total_len + tx_signature_length + tx_signatures[i] + "01" + from_pubkey_length + str(pubkey_list[i])
            print("input_signature:" + str(input_signature))
            raw_input = str(prev_hash_list[i]) + str(output_index_list[i]) + str(input_signature) + 'ffffffff'
            raw_inputs += raw_input

            inputs.append({"prev_hash": prev_hash_list[i],
                           "output_index": output_index_list[i],
                           "script": tx_signature_length + tx_signatures[i] + "01" + from_pubkey_length + str(pubkey_list[i]),
                           # "sign": tx_signatures[i] + "01",
                           "output_value": in_value_list[i],
                           "sequence": 4294967295,
                           "addressed": [in_address_list[i]], # 默认每个输入仅包含一个地址
                           "script_type": "pay-to-pubkey-hash",
                           })


        # 根据total、fee和输出数量随机生成输出
        out_address_or_opreturn = []
        sum_value = total - fees
        for i in range(out_num):
            if i < out_num-1:
                value = random.randint(1, sum_value - out_num + i + 1)
            else:
                value = sum_value
            out_address_or_opreturn.append((1, str(privkey_to_address(geneRandomKey())), value))
            sum_value = sum_value - value

        print('sum_value:', sum_value)

        # 计算不重复的输出地址
        for output in out_address_or_opreturn:
            if output[0] == 1:
                flag = True
                for address in addresses:
                    if output[1] == address:
                        flag = False
                if flag:
                    addresses.append(output[1])

        # 计算output
        output_dic = getOutputs(out_address_or_opreturn)

        # 计算raw_tx
        raw_tx = output_dic["version"] + str(raw_inputs) + output_dic["output_num"] + output_dic["output_script"] + '00000000'
        size = int(len(raw_tx) / 2)
        txid = big_small_end_convert(
            hashlib.sha256(hashlib.sha256(safe_from_hex(raw_tx)).digest()).hexdigest()).decode()
        # print("txid:" + str(txid))

        return {"hash": str(txid),
                "addresses": addresses,
                "total": total - fees,
                "fees": fees,
                "size": size,
                "vsize": size,
                "preference": "low",
                "vin_sz": in_sk_num,
                "vout_sz": len(out_address_or_opreturn),
                "data_protocol": "unknown",
                "inputs": inputs,
                "outputs": output_dic["outputs"], }

    def getOutputs(out_address_or_opreturn: List[Tuple[int, str, int]]) -> Dict:
        version = '01000000'
        # print("version:" + str(version))

        output_num = len(out_address_or_opreturn)
        output_num = add_to_len(big_small_end_convert(to_even(hex(output_num)[2:]).encode()), 2).decode()
        # print("output_num:" + str(output_num))

        output_script_list = list()
        outputs = list()

        for i in range(len(out_address_or_opreturn)):
            op_value = add_to_len(big_small_end_convert(to_even(hex(out_address_or_opreturn[i][2])[2:]).encode()), 16).decode()
            op_script = None
            if out_address_or_opreturn[i][0] == 0:  # 输出是opreturn
                message = out_address_or_opreturn[i][1]
                message_len = len(message)
                # print(message_len)
                if isinstance(message, bytes):
                    message_hex = message.hex()
                else:
                    if message_len > 0xff:
                        message_hex = message
                        message_len = message_len // 2
                    else:
                        message_hex = message.encode('ascii').hex()

                script_command = {255: [2, '4c'], 0xffff: [4, '4d'], 0xffffffff: [8, '4e']}
                for l, command in script_command.items():
                    if message_len <= l:
                        message_len = format(message_len, 'x').rjust(command[0], '0')
                        op_script_without_len = "6a" + command[1] + message_len + message_hex
                        op_len = to_even(hex(int(len(op_script_without_len) / 2))[2:])
                        op_script = op_len + op_script_without_len
                        break
                outputs.append({"value": out_address_or_opreturn[i][2],
                                "script": op_script[2:],
                                "addresses": None,
                                "script_type": "null-data",
                                "data_hex": message_hex,
                                "data_string": message, })
            elif out_address_or_opreturn[i][0] == 1:  # 输出是转账地址
                op_script = base58decode_to_P2PKH(out_address_or_opreturn[i][1])
                print("op_script:" + str(op_script))
                outputs.append({"value": out_address_or_opreturn[i][2],
                                "script": op_script[2:],
                                "addresses": [out_address_or_opreturn[i][1]],
                                "script_type": "pay-to-pubkey-hash", })
            output_script_list.append(op_value + op_script)

        # 测试时注释掉，实际生成数据时取消注释
        # random.shuffle(output_script_list)
        # print("output_script_list:" + str(output_script_list))
        output_script = ''
        for os in output_script_list:
            output_script += os
        # print("output_script:" + str(output_script))

        # output = str(output_num) + str(output_script)
        # print("output:" + str(output))

        return {"version": str(version),
                "output_num": str(output_num),
                "output_script_list": str(output_script_list),
                "output_script": str(output_script),
                "outputs": outputs, }  # 字典形式

    def getTosignTx(in_sk: int, prev_hash: str, output_index: int, out_address_or_opreturn: List[Tuple[int, str, int]])\
            -> Tuple[str, Dict]:
        version = '01000000'
        # print("version:" + str(version))
        input_num = 1
        input_num = add_to_len(big_small_end_convert(to_even(hex(input_num)[2:]).encode()), 2).decode()
        # print("input_num:" + str(input_num))
        prev_hash = big_small_end_convert(prev_hash).decode()
        # print("prev_hash:" + prev_hash)
        prev_hash_index = add_to_len(big_small_end_convert(to_even(hex(output_index)[2:]).encode()), 8).decode()
        # print("prev_hash_index:" + str(prev_hash_index))

        # 主链
        # input_pkhash = base58decode_to_P2PKH(privkey_to_address(in_sk))
        # 测试链
        input_pkhash = base58decode_to_P2PKH(privkey_to_address(in_sk, "111"))
        # print("input_pkhash:" + str(input_pkhash))

        sequence = 'ffffffff'
        # print("sequence:" + str(sequence))

        input = str(input_num) + prev_hash + str(prev_hash_index) + str(input_pkhash) + sequence
        # print("input:" + str(input))

        output_num = len(out_address_or_opreturn)
        output_num = add_to_len(big_small_end_convert(to_even(hex(output_num)[2:]).encode()), 2).decode()
        # print("output_num:" + str(output_num))

        output_script_list = list()
        outputs = list()

        for i in range(len(out_address_or_opreturn)):
            op_value = add_to_len(big_small_end_convert(to_even(hex(out_address_or_opreturn[i][2])[2:]).encode()), 16).decode()
            op_script = None
            if out_address_or_opreturn[i][0] == 0:  # 输出是opreturn
                message = out_address_or_opreturn[i][1]
                message_len = len(message)
                # print(message_len)
                if isinstance(message, bytes):
                    message_hex = message.hex()
                else:
                    if message_len > 0xff:
                        message_hex = message
                        message_len = message_len // 2
                    else:
                        message_hex = message.encode('ascii').hex()

                script_command = {255: [2, '4c'], 0xffff: [4, '4d'], 0xffffffff: [8, '4e']}
                for l, command in script_command.items():
                    if message_len <= l:
                        message_len = format(message_len, 'x').rjust(command[0], '0')
                        op_script_without_len = "6a" + command[1] + message_len + message_hex
                        op_len = to_even(hex(int(len(op_script_without_len) / 2))[2:])
                        op_script = op_len + op_script_without_len
                        break
                outputs.append({"value": out_address_or_opreturn[i][2],
                                "script": op_script[2:],
                                "addresses": None,
                                "script_type": "null-data",
                                "data_hex": message_hex,
                                "data_string": message, })
            elif out_address_or_opreturn[i][0] == 1:  # 输出是转账地址
                print("out_address_or_opreturn[i][1]:", out_address_or_opreturn[i][1])
                op_script = base58decode_to_P2PKH(out_address_or_opreturn[i][1])
                print("op_script:", op_script)
                # print("op_script:" + str(op_script))
                outputs.append({"value": out_address_or_opreturn[i][2],
                                "script": op_script[2:],
                                "addresses": [out_address_or_opreturn[i][1]],
                                "script_type": "pay-to-pubkey-hash", })
            output_script_list.append(op_value + op_script)

        # 测试时注释掉，实际生成数据时取消注释
        # random.shuffle(output_script_list)
        # print("output_script_list:" + str(output_script_list))
        output_script = ''
        for os in output_script_list:
            output_script += os
        # print("output_script:" + str(output_script))

        output = str(output_num) + str(output_script)
        # print("output:" + str(output))

        lock_time = '00000000'
        hash_type = '01000000'

        tosign_tx = version + input + output + lock_time + hash_type
        # print("tosign_tx:" + str(tosign_tx))

        return tosign_tx, {"version": str(version),
                           "input_num": str(input_num),
                           "prev_hash": str(prev_hash),
                           "prev_hash_index": str(prev_hash_index),
                           "input_pkhash": str(input_pkhash),
                           "sequence": str(sequence),
                           "output_num": str(output_num),
                           "output_script_list": str(output_script_list),
                           "output_script": str(output_script),
                           "outputs": outputs, }

    def base58decode_to_P2PKH(tmp: str) -> str:
        base58 = "123456789ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz"
        temp = []
        for i in tmp:
            temp.append(base58.index(i))
        tmp = temp[0]
        for i in range(len(temp) - 1):
            tmp = tmp * 58 + temp[i + 1]
        # print(len(hex(tmp)))
        print("tmp:", hex(tmp))
        if len(hex(tmp)) == 51:
            return '1976a9140' + hex(tmp)[4: -8] + '88ac'
        if len(hex(tmp)) == 52: # 测试链，6f开头
            return '1976a914' + hex(tmp)[4: -8] + '88ac'
        if len(hex(tmp)) == 49:
            return '1976a9140' + hex(tmp)[2: -8] + '88ac'
        return '1976a914' + hex(tmp)[2: -8] + '88ac'

    def make_tx_signatures_without_k(txs_to_sign, privkey_list, pubkey_list):
        """
        Loops through txs_to_sign and makes signatures using privkey_list and pubkey_list

        Not sure what privkeys and pubkeys to supply?
        Use get_input_addresses() to return a list of addresses.
        Matching those addresses to keys is up to you and how you store your private keys.
        A future version of this library may handle this for you, but it is not trivial.

        Note that if spending multisig funds the process is significantly more complicated.
        Each tx_to_sign must be signed by *each* private key.
        In a 2-of-3 transaction, two of [privkey1, privkey2, privkey3] must sign each tx_to_sign

        http://dev.blockcypher.com/#multisig-transactions
        """
        assert len(privkey_list) == len(pubkey_list) == len(txs_to_sign)
        # in the event of multiple inputs using the same pub/privkey,
        # that privkey should be included multiple times

        signatures = []
        for cnt, tx_to_sign in enumerate(txs_to_sign):
            sig = der_encode_sig(*ecdsa_raw_sign_without_k(tx_to_sign.rstrip(' \t\r\n\0'), privkey_list[cnt]))
            err_msg = 'Bad Signature: sig %s for tx %s with pubkey %s' % (
                sig,
                tx_to_sign,
                pubkey_list[cnt],
            )
            assert ecdsa_raw_verify(tx_to_sign, der_decode_sig(sig), pubkey_list[cnt]), err_msg
            signatures.append(sig)
        return signatures

    def ecdsa_raw_sign_without_k(msghash, priv):
        logger.info("msghash:" + msghash)
        z = hash_to_int(msghash)
        logger.info(z)
        k = deterministic_generate_k(msghash, priv)

        # k = 2 * k + 1  # 这里等式右边的k应该要限制不超过254比特，所以左边的k不超过255比特，肯定小于N，不需要对N求余
        # k = 10000000000
        r, y = fast_multiply(G, k)
        s = inv(k, N) * (z + r * decode_privkey(priv)) % N

        return 27 + ((y % 2) ^ (0 if s * 2 < N else 1)), r, s if s * 2 < N else N - s

    # print(base58decode_to_P2PKH("122Lfeq2Nmpzsg62hepT6zGixgX7YjSuRC"))
    # transfer
    # outputs = [(1, "n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo", 10), (1, "mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX", 3378771)]
    # random.shuffle(outputs)
    # for op in outputs:

    def geneRandomKey():
        sk = int(random_key(), 16)
        while len(base58decode_to_P2PKH(privkey_to_address(sk))) != 52:
            # print(len(base58decode_to_P2PKH(privkey_to_address(sk))))
            # print(sk)
            # print(base58decode_to_P2PKH(privkey_to_address(sk)))
            sk = int(random_key(), 16)
        # print('ok')
        return sk

    def geneTransferAmount():
        return random.randint(100000, 999999)

    def geneTotalAmount():
        return random.randint(10000000, 99999999)

    def geneFee():
        return random.randint(7800, 8600)

    def geneIndex():
        return random.randint(0, 1)

    def write_excel(data):
        wb = load_workbook("C://Users//ChenZhuo//Desktop//模拟隐蔽交易//test.xlsx")
        ws = wb.active
        # ws.append([1, 2, 3, 4, 5, 6])
        ws.append(data)
        wb.save("C://Users//ChenZhuo//Desktop//模拟隐蔽交易//test.xlsx")

    def write_json(tx):
        output_dir = "C://Users//ChenZhuo//Desktop//模拟隐蔽交易"
        listdir = os.listdir(output_dir)
        if 'test.json' in listdir:
            fr = open(pjoin(output_dir, 'test.json'), 'a')
            model = json.dumps(tx)
            fr.write(model + '\n')
            fr.close()

    def geneStrRandomHex64():
        res1 = str(uuid.uuid4())
        res1 = res1.replace('-', '')
        res2 = str(uuid.uuid4())
        res2 = res2.replace('-', '')
        return res1 + res2

    # print(base58decode_to_P2PKH("1EGyXLuxz4VAmJ5NJG8seQtteno4mfAzHA"))

    # # 第一轮数据生成测试
    # input_sk_list = list()
    # total_list = list()
    # prev_hash_list = list()
    # prev_index_list = list()
    #
    # init_sk = geneRandomKey()
    # input_sk_list.append(init_sk)
    # print(init_sk)
    #
    # init_total = geneTotalAmount()
    # total_list.append(init_total)
    # print(init_total)
    #
    # init_prev_hash = random_key()
    # prev_hash_list.append(init_prev_hash)
    # print(init_prev_hash)
    #
    # init_prev_index = geneIndex()
    # prev_index_list.append(init_prev_index)
    # print(init_prev_index)
    #
    # msg = "Everyones heart have a hero,maybe a person who save the people,maybe a animal who have got good spirit. So far,I cant forget a person who helped plenty of people. His name is Leifeng.I think he was a good person I cant understand why he was so friendly,but I really know,he was a hero in my heart.I looked up upon him.and I will do all things I can do,I want to be the 2nd Leifeng. When I was young he would kiss me every night and sing to me as I awoke. He made me feel so special. I knew that my father loved me. It made me feel strong enough to do almost everything. He got sick, but strength(力气) kept him alive. He loved his children so much. He gave up years of his life caring for the woman, his wife. She was going to die. Day and night he struggled(努力) for years with her disease. （疾病）Despite disease of his own, he still stood by her side caring for her and loving her until the day she died. Twenty days later, he died. I lost my hero, my father, a man who was my friend. A few years have passed, and the life has changed a lot. But the love for my father won’t change and it will never end."
    # c_msg = encrypt(msg).decode()
    # b_c_msg = bin(int(c_msg, 16))[2:]
    #
    # for i in range(int(len(b_c_msg) / 254)):
    #     k = int(b_c_msg[i * 254: (i + 1) * 254], 2)
    #     transfer = geneTransferAmount()
    #     fee = geneFee()
    #     if total_list[i] - fee - transfer < 0:
    #         print("Amount is not enough:", i)
    #         break
    #
    #     sk = geneRandomKey()
    #     input_sk_list.append(sk)
    #     total_list.append(total_list[i] - fee - transfer)
    #     index = geneIndex()
    #     prev_index_list.append(index)
    #     output = list()
    #     if index == 0:
    #         output.append((1, privkey_to_address(sk), total_list[i] - fee - transfer))
    #         output.append((1, privkey_to_address(geneRandomKey()), transfer))
    #     elif index == 1:
    #         output.append((1, privkey_to_address(geneRandomKey()), transfer))
    #         output.append((1, privkey_to_address(sk), total_list[i] - fee - transfer))
    #
    #     tx_info = getSingleTx(input_sk_list[i], prev_hash_list[i], prev_index_list[i], output, total_list[i], fee, k)
    #     write_excel([1, 2, 2, 1.5, total_list[i] - fee, fee])
    #     write_json(tx_info)
    #
    #     prev_hash_list.append(tx_info["hash"])
    #     i += 1

    # 模拟生成单输入交易测试
    # # 无op_return
    # print(getSingleTxWithoutK258(107154816688509022565992420189818653035020061696220201248865645576915755526209,
    #                   "31785680464a3b3c4a2b113f1d36f40c9289d5aec77939425e73370278fdc831", 2,
    #                   [(1, "n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo", 10),
    #                    (1, "mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX", 3378771)], 3386581, 7800))
    # print(getSingleTxWithoutK226(107154816688509022565992420189818653035020061696220201248865645576915755526209,
    #                   "31785680464a3b3c4a2b113f1d36f40c9289d5aec77939425e73370278fdc831", 2,
    #                   [(1, "n3J35FPrrDtrZiRasVki6J4r6JbZMgexzo", 10),
    #                    (1, "mxrNg8AsWmGJcjFBsJwHswXnDk1z261bfX", 3378771)], 3386581, 7800))
    #
    # # 有op_return
    # print(getSingleTx(106293531705061254553737748519143550965015022890347762128114678278566473882258,
    #                   "c6cfc79fae7ae008e2d2bed8fd010eae3a9c003906ea563ce83d8acd85af49b8", 1,
    #                   [(0, "test", 0), (1, "mgntfnt9aHYD1xFUD6KNBVniY6rMwbWVum", 577076)],
    #                   584476, 7400))

    # 模拟生成m输入n输出交易测试
    # print(getSingleMultiInputTx(20, 12, 12345, 123))

    # 根据txt文件生成m输入n输出的交易

    # # 第二轮数据生成测试（多输入多输出）
    # 读取excel
    # 得到工作簿对象
    # import openpyxl
    # workbook = openpyxl.load_workbook("C://Users//ChenZhuo//Desktop//模拟隐蔽交易//gen_data.xlsx")
    # # 获取当前活动的sheet页对象
    # sheet = workbook.active
    # # 获取指定的sheet页对象
    # sheet1 = workbook["Sheet1"]

    # # 第二种读取方式: _get_cell(row,cloumn)
    # print(sheet._get_cell(1, 1).value)

    # for i in range(2, 991):
    #     # print(i, int(sheet._get_cell(i, 3).value), int(sheet._get_cell(i, 4).value), int(sheet._get_cell(i, 1).value), int(sheet._get_cell(i, 2).value))
    #     total = int(sheet._get_cell(i, 1).value)
    #     fees = int(sheet._get_cell(i, 2).value)
    #     value = int(sheet._get_cell(i, 3).value)
    #     print(i, total, fees, value)
    #     # txinfo = getSingleMultiInputTx(int(sheet._get_cell(i, 3).value), int(sheet._get_cell(i, 4).value), int(sheet._get_cell(i, 1).value), int(sheet._get_cell(i, 2).value))
    #     # write_json(txinfo)
    #
    # # 如果只读取了数据可以直接关闭
    # workbook.close()


    # 第三轮数据生成测试（全部1输入2输出）
    input_sk_list = list()
    total_list = list()
    prev_hash_list = list()
    prev_index_list = list()

    # init_sk = geneRandomKey()
    # input_sk_list.append(init_sk)
    # print(init_sk)

    # init_total = geneTotalAmount()
    # total_list.append(init_total)
    # print(init_total)

    # init_prev_hash = random_key()
    # prev_hash_list.append(init_prev_hash)
    # print(init_prev_hash)

    # init_prev_index = geneIndex()
    # prev_index_list.append(init_prev_index)
    # print(init_prev_index)

    # msg = "Everyones heart have a hero,maybe a person who save the people,maybe a animal who have got good spirit. So far,I cant forget a person who helped plenty of people. His name is Leifeng.I think he was a good person I cant understand why he was so friendly,but I really know,he was a hero in my heart.I looked up upon him.and I will do all things I can do,I want to be the 2nd Leifeng. When I was young he would kiss me every night and sing to me as I awoke. He made me feel so special. I knew that my father loved me. It made me feel strong enough to do almost everything. He got sick, but strength(力气) kept him alive. He loved his children so much. He gave up years of his life caring for the woman, his wife. She was going to die. Day and night he struggled(努力) for years with her disease. （疾病）Despite disease of his own, he still stood by her side caring for her and loving her until the day she died. Twenty days later, he died. I lost my hero, my father, a man who was my friend. A few years have passed, and the life has changed a lot. But the love for my father won’t change and it will never end."
    # c_msg = encrypt(msg).decode()
    # b_c_msg = bin(int(c_msg, 16))[2:]

    import openpyxl
    workbook = openpyxl.load_workbook("C://Users//ChenZhuo//Desktop//模拟隐蔽交易//gen_data.xlsx")
    # 获取当前活动的sheet页对象
    sheet = workbook.active
    # 获取指定的sheet页对象
    sheet1 = workbook["Sheet1"]

    for i in range(951, 991):
        total = int(sheet._get_cell(i, 1).value)
        fees = int(sheet._get_cell(i, 2).value)
        value = int(sheet._get_cell(i, 3).value)
        if total - fees - value <= 0:
            print("Amount is not enough:", i)
            continue
        print(i, total, fees, value)

        # transfer = geneTransferAmount()
        # fee = geneFee()
        # if total_list[i] - fee - transfer < 0:
        #     print("Amount is not enough:", i)
        #     break

        sk = geneRandomKey()
        # input_sk_list.append(sk)
        # total_list.append(total)
        index = geneIndex()
        # prev_index_list.append(index)
        output = list()
        if index == 0:
            output.append((1, privkey_to_address(geneRandomKey()), total - fees - value))
            output.append((1, privkey_to_address(geneRandomKey()), value))
        elif index == 1:
            output.append((1, privkey_to_address(geneRandomKey()), value))
            output.append((1, privkey_to_address(geneRandomKey()), total - fees - value))

        tx_info = getSingleTxWithoutK226(sk, random_key(), index, output, total, fees)
        # write_excel([1, 2, 2, 1.5, total_list[i] - fee, fee])
        write_json(tx_info)

        # prev_hash_list.append(tx_info["hash"])

    workbook.close()